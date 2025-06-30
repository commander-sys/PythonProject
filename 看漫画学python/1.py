import openpyxl as ol
from openpyxl.reader.excel import load_workbook

file1 =  ol.load_workbook('DCGW.xlsx')

sheet = file1['Sheet1']

rows = sheet.max_row

print(rows)